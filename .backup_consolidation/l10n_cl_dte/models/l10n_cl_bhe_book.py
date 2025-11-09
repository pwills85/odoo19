# -*- coding: utf-8 -*-
"""
Libro de Boletas de Honorarios Electrónicas
Según normativa SII - Resolución Exenta N° 34 del 2019

Compliance SII:
- Registro mensual obligatorio de BHE recibidas
- Declaración en F29 mensual
- Formato Excel según plantilla SII
- Campos obligatorios según Art. 74 N°5 Ley de la Renta
"""

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
from dateutil.relativedelta import relativedelta
import logging
import base64
from io import BytesIO

_logger = logging.getLogger(__name__)


class L10nClBheBook(models.Model):
    """
    Libro de Boletas de Honorarios Electrónicas
    Monthly book for tax reporting of BHE documents

    Según SII:
    - Obligatorio para empresas que reciben BHE
    - Debe generarse mensualmente
    - Información se declara en F29
    - Formato: Excel con columnas específicas SII
    """
    _name = "l10n_cl.bhe.book"
    _description = "Libro de Boletas de Honorarios"
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = "period_year desc, period_month desc"

    # ═══════════════════════════════════════════════════════════
    # CAMPOS BÁSICOS
    # ═══════════════════════════════════════════════════════════

    name = fields.Char(
        string="Nombre",
        compute="_compute_name",
        store=True
    )

    display_name = fields.Char(
        string="Display Name",
        compute="_compute_display_name",
        store=True
    )

    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        required=True,
        default=lambda self: self.env.company
    )

    currency_id = fields.Many2one(
        'res.currency',
        string='Moneda',
        default=lambda self: self.env.company.currency_id
    )

    # ═══════════════════════════════════════════════════════════
    # PERÍODO (Según SII - Mensual)
    # ═══════════════════════════════════════════════════════════

    period_year = fields.Integer(
        string="Año",
        required=True,
        default=lambda self: fields.Date.today().year,
        tracking=True
    )

    period_month = fields.Selection([
        ('1', 'Enero'),
        ('2', 'Febrero'),
        ('3', 'Marzo'),
        ('4', 'Abril'),
        ('5', 'Mayo'),
        ('6', 'Junio'),
        ('7', 'Julio'),
        ('8', 'Agosto'),
        ('9', 'Septiembre'),
        ('10', 'Octubre'),
        ('11', 'Noviembre'),
        ('12', 'Diciembre'),
    ], string="Mes", required=True,
    default=lambda self: str(fields.Date.today().month),
    tracking=True)

    date_from = fields.Date(
        string="Desde",
        compute="_compute_dates",
        store=True,
        help="Primer día del mes"
    )

    date_to = fields.Date(
        string="Hasta",
        compute="_compute_dates",
        store=True,
        help="Último día del mes"
    )

    # ═══════════════════════════════════════════════════════════
    # LÍNEAS DEL LIBRO (Según SII)
    # ═══════════════════════════════════════════════════════════

    line_ids = fields.One2many(
        'l10n_cl.bhe.book.line',
        'book_id',
        string="Líneas del Libro",
        help="Detalle de BHE recibidas en el período"
    )

    # ═══════════════════════════════════════════════════════════
    # TOTALES (Según F29 - Declaración Jurada)
    # ═══════════════════════════════════════════════════════════

    total_count = fields.Integer(
        string="Total BHE",
        compute="_compute_totals",
        store=True,
        help="Cantidad total de BHE recibidas"
    )

    total_gross = fields.Monetary(
        string="Total Monto Bruto",
        compute="_compute_totals",
        store=True,
        currency_field='currency_id',
        help="Suma total de honorarios brutos (antes retención)"
    )

    total_retention = fields.Monetary(
        string="Total Retenciones",
        compute="_compute_totals",
        store=True,
        currency_field='currency_id',
        help="Suma total de retenciones efectuadas. Se declara en F29 línea 150"
    )

    total_net = fields.Monetary(
        string="Total Neto Pagado",
        compute="_compute_totals",
        store=True,
        currency_field='currency_id',
        help="Suma total pagada a profesionales (bruto - retención)"
    )

    # ═══════════════════════════════════════════════════════════
    # ESTADO
    # ═══════════════════════════════════════════════════════════

    state = fields.Selection([
        ('draft', 'Borrador'),
        ('posted', 'Confirmado'),
        ('declared', 'Declarado en F29'),
        ('sent', 'Enviado al SII')
    ], string='Estado', default='draft', tracking=True, copy=False)

    # ═══════════════════════════════════════════════════════════
    # DECLARACIÓN F29 (Integración)
    # ═══════════════════════════════════════════════════════════

    f29_declaration_date = fields.Date(
        string="Fecha Declaración F29",
        readonly=True,
        help="Fecha en que se declaró este libro en el F29"
    )

    f29_line_150 = fields.Monetary(
        string="F29 Línea 150 (Retenciones)",
        compute="_compute_f29_line_150",
        store=True,
        currency_field='currency_id',
        help="Monto a declarar en F29 línea 150 (Retenciones Art. 42 N°2)"
    )

    # ═══════════════════════════════════════════════════════════
    # EXPORTACIÓN (Según formato SII)
    # ═══════════════════════════════════════════════════════════

    export_file = fields.Binary(
        string="Archivo Excel SII",
        attachment=True,
        help="Libro exportado en formato Excel según plantilla SII"
    )

    export_filename = fields.Char(
        string="Nombre Archivo",
        compute="_compute_export_filename",
        help="Formato: LibroBHE_YYYYMM_RUT.xlsx"
    )

    # ═══════════════════════════════════════════════════════════
    # NOTAS
    # ═══════════════════════════════════════════════════════════

    notes = fields.Text(
        string="Notas",
        help="Observaciones sobre el libro mensual"
    )

    # ═══════════════════════════════════════════════════════════
    # COMPUTED FIELDS
    # ═══════════════════════════════════════════════════════════

    @api.depends('period_year', 'period_month')
    def _compute_name(self):
        for rec in self:
            if rec.period_year and rec.period_month:
                month_name = dict(rec._fields['period_month'].selection)[rec.period_month]
                rec.name = f"Libro BHE {month_name} {rec.period_year}"
            else:
                rec.name = "Libro BHE"

    @api.depends('name')
    def _compute_display_name(self):
        for rec in self:
            rec.display_name = rec.name

    @api.depends('period_year', 'period_month')
    def _compute_dates(self):
        for rec in self:
            if rec.period_year and rec.period_month:
                month = int(rec.period_month)
                date_from = fields.Date(rec.period_year, month, 1)
                date_to = date_from + relativedelta(day=31)
                rec.date_from = date_from
                rec.date_to = date_to
            else:
                rec.date_from = False
                rec.date_to = False

    @api.depends('line_ids', 'line_ids.amount_gross', 'line_ids.amount_retention', 'line_ids.amount_net')
    def _compute_totals(self):
        for rec in self:
            rec.total_count = len(rec.line_ids)
            rec.total_gross = sum(rec.line_ids.mapped('amount_gross'))
            rec.total_retention = sum(rec.line_ids.mapped('amount_retention'))
            rec.total_net = sum(rec.line_ids.mapped('amount_net'))

    @api.depends('total_retention')
    def _compute_f29_line_150(self):
        """
        Calcula monto a declarar en F29 línea 150.

        Según SII:
        - Línea 150: Retenciones Art. 42 N°2 (Honorarios)
        - Corresponde al total de retenciones efectuadas en el mes
        """
        for rec in self:
            rec.f29_line_150 = rec.total_retention

    @api.depends('period_year', 'period_month', 'company_id')
    def _compute_export_filename(self):
        """
        Genera nombre archivo según convención SII.

        Formato: LibroBHE_YYYYMM_RUT.xlsx
        Ejemplo: LibroBHE_202501_76123456-7.xlsx
        """
        for rec in self:
            if rec.period_year and rec.period_month and rec.company_id:
                year_month = f"{rec.period_year}{rec.period_month.zfill(2)}"
                rut = rec.company_id.vat or "99999999-9"
                rut_clean = rut.replace('.', '').replace('-', '')
                rec.export_filename = f"LibroBHE_{year_month}_{rut_clean}.xlsx"
            else:
                rec.export_filename = "LibroBHE.xlsx"

    # ═══════════════════════════════════════════════════════════
    # CONSTRAINTS
    # ═══════════════════════════════════════════════════════════

    @api.constrains('period_year', 'period_month')
    def _check_period(self):
        for rec in self:
            if rec.period_year < 2018 or rec.period_year > 2099:
                raise ValidationError(
                    _("El año debe estar entre 2018 y 2099.")
                )

    # CONSTRAINTS (Odoo 19 CE format)
    _period_unique = models.Constraint(
        'unique(period_year, period_month, company_id)',
        'Ya existe un Libro BHE para este período en esta compañía.'
    )

    # ═══════════════════════════════════════════════════════════
    # ACTIONS
    # ═══════════════════════════════════════════════════════════

    def action_generate_lines(self):
        """
        Genera líneas desde BHE del período.

        Según SII:
        - Solo BHE contabilizadas (state = posted o accepted)
        - Ordenadas por fecha y número
        - Incluye todas las BHE recibidas en el mes
        """
        for rec in self:
            if rec.state not in ('draft', 'posted'):
                raise UserError(
                    _("Solo se pueden generar líneas en estado Borrador o Confirmado.")
                )

            # Buscar BHE del período
            bhes = self.env['l10n_cl.bhe'].search([
                ('company_id', '=', rec.company_id.id),
                ('date', '>=', rec.date_from),
                ('date', '<=', rec.date_to),
                ('state', 'in', ['posted', 'accepted'])
            ], order='date asc, number asc')

            if not bhes:
                raise UserError(
                    _(f"No se encontraron BHE contabilizadas en el período "
                      f"{rec.period_month}/{rec.period_year}.\n\n"
                      f"Contabilice primero las BHE recibidas antes de generar el libro.")
                )

            # Limpiar líneas existentes
            rec.line_ids.unlink()

            # Crear líneas según orden SII (fecha + número)
            line_number = 1
            for bhe in bhes:
                self.env['l10n_cl.bhe.book.line'].create({
                    'book_id': rec.id,
                    'line_number': line_number,
                    'bhe_id': bhe.id,
                    'bhe_date': bhe.date,
                    'bhe_number': bhe.number,
                    'partner_id': bhe.partner_id.id,
                    'partner_vat': bhe.partner_vat,
                    'partner_name': bhe.partner_id.name,
                    'service_description': bhe.service_description,
                    'amount_gross': bhe.amount_gross,
                    'retention_rate': bhe.retention_rate,
                    'amount_retention': bhe.amount_retention,
                    'amount_net': bhe.amount_net,
                })
                line_number += 1

            _logger.info(
                f"✅ Libro BHE {rec.name}: {len(bhes)} BHE procesadas"
            )

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Libro Generado'),
                    'message': _(f'Se generaron {len(bhes)} líneas desde BHE contabilizadas.'),
                    'type': 'success',
                    'sticky': False,
                }
            }

    def action_post(self):
        """
        Confirmar libro.

        Según SII:
        - Libro debe estar completo con todas las BHE del mes
        - Una vez confirmado, no se pueden agregar más líneas
        - Habilita exportación a Excel
        """
        for rec in self:
            if not rec.line_ids:
                raise UserError(
                    _("El libro no tiene líneas.\n\n"
                      "Use el botón 'Generar Líneas' para cargar las BHE del período.")
                )

            if rec.total_retention == 0:
                raise UserError(
                    _("El libro tiene retenciones en $0.\n\n"
                      "Verifique que las BHE tengan montos correctos.")
                )

            rec.write({'state': 'posted'})

            _logger.info(
                f"✅ Libro BHE confirmado: {rec.name} "
                f"({rec.total_count} BHE, ${rec.total_retention:,.0f} retenciones)"
            )

    def action_mark_declared_f29(self):
        """
        Marca libro como declarado en F29.

        Según SII:
        - Se ejecuta DESPUÉS de presentar F29 en el SII
        - Registra fecha de declaración
        - Bloquea modificaciones
        """
        for rec in self:
            if rec.state != 'posted':
                raise UserError(_("Solo se pueden marcar como declarados libros confirmados."))

            rec.write({
                'state': 'declared',
                'f29_declaration_date': fields.Date.today()
            })

            _logger.info(f"✅ Libro BHE declarado en F29: {rec.name}")

    def action_export_excel(self):
        """
        Exporta libro a Excel según formato SII.

        Formato según SII:
        Columnas obligatorias:
        1. N° Correlativo
        2. Fecha BHE
        3. N° BHE
        4. RUT Prestador
        5. Nombre Prestador
        6. Descripción Servicio
        7. Monto Bruto
        8. Tasa Retención (%)
        9. Monto Retención
        10. Monto Neto Pagado
        """
        self.ensure_one()

        if not self.line_ids:
            raise UserError(
                _("El libro no tiene líneas para exportar.\n\n"
                  "Use 'Generar Líneas' primero.")
            )

        try:
            # Importar openpyxl (debe estar en requirements)
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise UserError(
                _("No se puede exportar a Excel.\n\n"
                  "Instale la librería openpyxl: pip install openpyxl")
            )

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Libro BHE {self.period_month}/{self.period_year}"

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título del libro
        ws['A1'] = f"LIBRO DE BOLETAS DE HONORARIOS ELECTRÓNICAS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Período: {dict(self._fields['period_month'].selection)[self.period_month]} {self.period_year}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"RUT Empresa: {self.company_id.vat}"
        ws['A4'] = f"Razón Social: {self.company_id.name}"
        ws['A5'] = f"Total Retenciones (F29 Línea 150): ${self.f29_line_150:,.0f}"
        ws['A5'].font = Font(bold=True)

        # Headers (fila 7)
        headers = [
            'N°',
            'Fecha BHE',
            'N° BHE',
            'RUT Prestador',
            'Nombre Prestador',
            'Descripción Servicio',
            'Monto Bruto',
            'Tasa Ret. (%)',
            'Monto Retención',
            'Monto Neto Pagado'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Datos (desde fila 8)
        row_num = 8
        for line in self.line_ids.sorted('line_number'):
            ws.cell(row=row_num, column=1).value = line.line_number
            ws.cell(row=row_num, column=2).value = line.bhe_date.strftime('%d/%m/%Y')
            ws.cell(row=row_num, column=3).value = line.bhe_number
            ws.cell(row=row_num, column=4).value = line.partner_vat
            ws.cell(row=row_num, column=5).value = line.partner_name
            ws.cell(row=row_num, column=6).value = line.service_description[:100]
            ws.cell(row=row_num, column=7).value = line.amount_gross
            ws.cell(row=row_num, column=8).value = line.retention_rate
            ws.cell(row=row_num, column=9).value = line.amount_retention
            ws.cell(row=row_num, column=10).value = line.amount_net

            # Formato moneda
            ws.cell(row=row_num, column=7).number_format = '#,##0'
            ws.cell(row=row_num, column=8).number_format = '0.00'
            ws.cell(row=row_num, column=9).number_format = '#,##0'
            ws.cell(row=row_num, column=10).number_format = '#,##0'

            # Borders
            for col in range(1, 11):
                ws.cell(row=row_num, column=col).border = border

            row_num += 1

        # Totales
        total_row = row_num
        ws.cell(row=total_row, column=6).value = "TOTALES:"
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=7).value = self.total_gross
        ws.cell(row=total_row, column=7).number_format = '#,##0'
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=9).value = self.total_retention
        ws.cell(row=total_row, column=9).number_format = '#,##0'
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        ws.cell(row=total_row, column=10).value = self.total_net
        ws.cell(row=total_row, column=10).number_format = '#,##0'
        ws.cell(row=total_row, column=10).font = Font(bold=True)

        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15

        # Guardar
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Actualizar registro
        self.write({
            'export_file': base64.b64encode(excel_file.read())
        })

        _logger.info(f"✅ Libro BHE exportado: {self.export_filename}")

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Exportación Exitosa'),
                'message': _(f'Libro exportado: {self.export_filename}'),
                'type': 'success',
                'sticky': False,
            }
        }

    def action_draft(self):
        """Volver a borrador"""
        for rec in self:
            if rec.state == 'declared':
                raise UserError(
                    _("No se puede volver a borrador un libro declarado en F29.\n\n"
                      "Contacte al contador si necesita corregir.")
                )

            rec.write({'state': 'draft'})


class L10nClBheBookLine(models.Model):
    """
    Línea de Libro de Boletas de Honorarios

    Según SII:
    - Cada línea representa una BHE recibida
    - Debe incluir todos los campos obligatorios
    - Se ordena por fecha y número correlativo
    """
    _name = "l10n_cl.bhe.book.line"
    _description = "Línea de Libro de Boletas de Honorarios"
    _order = "book_id, line_number"

    # ═══════════════════════════════════════════════════════════
    # RELACIÓN CON LIBRO
    # ═══════════════════════════════════════════════════════════

    book_id = fields.Many2one(
        'l10n_cl.bhe.book',
        string="Libro",
        required=True,
        ondelete='cascade',
        index=True
    )

    line_number = fields.Integer(
        string="N° Línea",
        required=True,
        help="Número correlativo en el libro"
    )

    # ═══════════════════════════════════════════════════════════
    # REFERENCIA A BHE ORIGINAL
    # ═══════════════════════════════════════════════════════════

    bhe_id = fields.Many2one(
        'l10n_cl.bhe',
        string="BHE",
        readonly=True,
        help="BHE original que genera esta línea"
    )

    # ═══════════════════════════════════════════════════════════
    # DATOS BHE (Según SII - Campos Obligatorios)
    # ═══════════════════════════════════════════════════════════

    bhe_date = fields.Date(
        string="Fecha BHE",
        required=True,
        help="Fecha de emisión de la BHE"
    )

    bhe_number = fields.Char(
        string="Número BHE",
        required=True,
        help="Número de folio de la BHE"
    )

    partner_id = fields.Many2one(
        'res.partner',
        string="Prestador",
        required=True,
        help="Profesional que emitió la BHE"
    )

    partner_vat = fields.Char(
        string="RUT Prestador",
        required=True,
        help="RUT del profesional (formato: 12.345.678-9)"
    )

    partner_name = fields.Char(
        string="Nombre Prestador",
        required=True,
        help="Nombre completo del profesional"
    )

    service_description = fields.Text(
        string="Descripción Servicio",
        help="Descripción de los servicios profesionales prestados"
    )

    # ═══════════════════════════════════════════════════════════
    # MONTOS (Según SII)
    # ═══════════════════════════════════════════════════════════

    currency_id = fields.Many2one(
        related='book_id.currency_id',
        string='Moneda',
        store=True
    )

    amount_gross = fields.Monetary(
        string="Monto Bruto",
        required=True,
        currency_field='currency_id',
        help="Monto total de honorarios antes de retención"
    )

    retention_rate = fields.Float(
        string="Tasa Retención (%)",
        required=True,
        digits=(5, 2),
        help="Tasa de retención aplicada (varía según año)"
    )

    amount_retention = fields.Monetary(
        string="Monto Retención",
        required=True,
        currency_field='currency_id',
        help="Monto retenido (Bruto × Tasa%)"
    )

    amount_net = fields.Monetary(
        string="Monto Neto Pagado",
        required=True,
        currency_field='currency_id',
        help="Monto pagado al profesional (Bruto - Retención)"
    )

    # ═══════════════════════════════════════════════════════════
    # DISPLAY
    # ═══════════════════════════════════════════════════════════

    display_name = fields.Char(
        string="Display Name",
        compute="_compute_display_name",
        store=True
    )

    @api.depends('line_number', 'bhe_number', 'partner_name')
    def _compute_display_name(self):
        for rec in self:
            if rec.line_number and rec.bhe_number and rec.partner_name:
                rec.display_name = f"Línea {rec.line_number}: BHE {rec.bhe_number} - {rec.partner_name}"
            else:
                rec.display_name = f"Línea {rec.line_number or 'N/A'}"
